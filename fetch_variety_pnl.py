#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
大盘手网「品种盈亏」每日采集：访问账户页，抓取图表数据，写入 Excel，并记录相对前一日的变化。

依赖：pip install -r requirements.txt && playwright install chromium

用法：
  python fetch_variety_pnl.py
  python fetch_variety_pnl.py --url https://www.dpswang.com/account/6-34668.html
  python fetch_variety_pnl.py --headed   # 调试时显示浏览器

登录（任选其一，推荐敏感信息用环境变量，勿把密码写进命令行历史）：
  set DPSWANG_USERNAME=你的手机号或账号
  set DPSWANG_PASSWORD=你的密码
  python fetch_variety_pnl.py --storage-state dpswang_auth.json --save-storage-state dpswang_auth.json

  或仅自动登录（每次启动都会先打开登录页）：
  python fetch_variety_pnl.py --username xxx --password yyy

  若站点有图形验证码/短信验证，自动登录会失败，请用 --headed 手动登录一次并配合
  --save-storage-state 写入会话文件，之后定时任务只带 --storage-state 即可无头运行。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# 默认账户页（赛季-账户）
DEFAULT_ACCOUNT_URL = "https://www.dpswang.com/account/7-{account_id}.html"
DEFAULT_LOGIN_URL = "https://www.dpswang.com/login"
DEFAULT_EXCEL_NAME = "品种盈亏.xlsx"
ENV_USERNAME = "DPSWANG_USERNAME"
ENV_PASSWORD = "DPSWANG_PASSWORD"
SHEET_DAILY_SUFFIX = "每日数据"
SHEET_DIFF_SUFFIX = "每日变化"

# 用户映射：key=用户名（作为 Sheet 名），value=账户编号
ACCOUNTS: dict[str, str] = {
    # "karin_hl": "39387",
    "猪珊珊": "34668",
    "鲁四老爷": "28994",
    "闪亮的脚趾": "38537",
}


def _script_dir() -> Path:
    return Path(__file__).resolve().parent


def _normalize_name(s: str) -> str:
    return str(s).strip()


def _parse_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
            if m:
                return float(m.group(0))
    return None


def _try_extract_from_json(obj: Any) -> dict[str, float] | None:
    """
    从多种可能的 JSON 结构中解析「品种名 -> 盈亏」。
    """
    if obj is None:
        return None

    if isinstance(obj, dict):
        # 常见：{ "data": [...] 或 "data": {...} }
        for key in ("data", "list", "rows", "items", "result", "records"):
            if key not in obj:
                continue
            inner = obj[key]
            if isinstance(inner, list):
                got = _try_extract_from_json(inner)
                if got:
                    return got
            elif isinstance(inner, dict):
                got = _try_extract_from_json(inner)
                if got:
                    return got

        # ECharts 配置风格: { xAxis: {data: [...]}, series: [{data: [...]}] }
        if "xAxis" in obj or "series" in obj:
            cats: list[str] = []
            xa = obj.get("xAxis")
            if isinstance(xa, list) and xa:
                xd = xa[0].get("data") if isinstance(xa[0], dict) else None
                if isinstance(xd, list):
                    cats = [_normalize_name(x) for x in xd]
            elif isinstance(xa, dict):
                xd = xa.get("data")
                if isinstance(xd, list):
                    cats = [_normalize_name(x) for x in xd]
            # 也检查 categories 字段
            if not cats:
                for ck in ("categories", "xData", "xLabels"):
                    cv = obj.get(ck)
                    if isinstance(cv, list) and cv:
                        cats = [_normalize_name(x) for x in cv]
                        break

            all_vals: list[list[float]] = []
            se = obj.get("series")
            if isinstance(se, list):
                for s_item in se:
                    if not isinstance(s_item, dict):
                        continue
                    d = s_item.get("data")
                    if not isinstance(d, list):
                        continue
                    row: list[float] = []
                    for item in d:
                        if isinstance(item, dict):
                            v = item.get("value")
                            if isinstance(v, (list, tuple)) and v:
                                row.append(float(v[0]))
                            else:
                                n = _parse_number(v)
                                row.append(n if n is not None else 0.0)
                        else:
                            n = _parse_number(item)
                            row.append(n if n is not None else 0.0)
                    if row:
                        all_vals.append(row)

            if cats and all_vals:
                merged: dict[str, float] = {}
                for row in all_vals:
                    for i, nm in enumerate(cats):
                        if i < len(row):
                            merged[nm] = merged.get(nm, 0.0) + row[i]
                if len(merged) >= 2:
                    return merged

        # 直接是 { "品种名": 数值, ... } 格式
        _FRAMEWORK_KEYS = frozenset({
            "data", "state", "once", "_errors", "serverRendered", "routePath",
            "config", "layout", "error", "fetch", "mutations", "modules",
            "status", "code", "msg", "message", "total", "success", "timestamp",
            "page", "limit", "offset", "count", "version", "env",
        })
        scalar_pairs: dict[str, float] = {}
        for k, v in obj.items():
            if isinstance(k, str) and isinstance(v, (int, float)) and not isinstance(v, bool):
                if k.lower() in _FRAMEWORK_KEYS:
                    continue
                scalar_pairs[k] = float(v)
        has_cjk = any(re.search(r"[\u4e00-\u9fff]", k) for k in scalar_pairs)
        if len(scalar_pairs) >= 3 and has_cjk:
            return scalar_pairs

    if isinstance(obj, list):
        out: dict[str, float] = {}
        for item in obj:
            if not isinstance(item, dict):
                continue
            # 尝试大量可能的字段名组合
            name = None
            for nk in ("varietyName", "variety_name", "name", "symbolName",
                       "productName", "product_name", "品种", "contract",
                       "variety", "symbol", "product", "instrument"):
                if nk in item and item[nk] is not None:
                    name = item[nk]
                    break
            val = None
            for vk in ("profit", "pnl", "value", "盈亏", "profitLoss",
                       "profit_loss", "amount", "totalProfit", "total_profit",
                       "netProfit", "net_profit", "money", "sum"):
                if vk in item and item[vk] is not None:
                    val = item[vk]
                    break
            if name is not None and val is not None:
                n = _parse_number(val)
                if n is not None:
                    out[_normalize_name(str(name))] = n
        if len(out) >= 2:
            return out

    return None


def _merge_dicts(candidates: list[dict[str, float] | None]) -> dict[str, float] | None:
    best: dict[str, float] = {}
    best_len = 0
    for d in candidates:
        if not d:
            continue
        if len(d) > best_len:
            best = dict(d)
            best_len = len(d)
    return best if best_len >= 2 else None


def _cjk_key_ratio(d: dict[str, float]) -> float:
    if not d:
        return 0.0
    n = sum(1 for k in d if re.search(r"[\u4e00-\u9fff]", str(k)))
    return n / len(d)


def _variety_map_quality_score(d: dict[str, float]) -> float:
    """分数越高越像「豆一、沪银」等中文品种名（避免 API 内部代码覆盖）。"""
    if not d or len(d) < 2:
        return -1.0
    return len(d) * 10.0 + _cjk_key_ratio(d) * 500.0


def _pick_best_variety_maps(
    sources: list[tuple[str, dict[str, float] | None, str]],
) -> dict[str, float] | None:
    """
    sources: (来源名, 数据, 附加说明，如 ECharts 标题)
    优先：标题含「品种盈亏」的 ECharts；其次中文键占比高且条数多的。
    """
    best: dict[str, float] | None = None
    best_score = -float("inf")
    for _src, data, hint in sources:
        if not data or len(data) < 2:
            continue
        score = _variety_map_quality_score(data)
        if "品种盈亏" in (hint or ""):
            score += 8000.0
        # 对 ECharts 来源加分（因为直接从图表读取最可靠）
        if _src == "echarts":
            score += 2000.0
        elif _src == "iframe":
            score += 1000.0
        print(f"  [评分] 来源={_src}, 品种数={len(data)}, 分数={score:.0f}, "
              f"中文比={_cjk_key_ratio(data):.0%}, 提示=\"{hint}\"")
        if score > best_score:
            best_score = score
            best = dict(data)
    return best


def _deep_search_variety_map(obj: Any, max_depth: int = 18, _depth: int = 0) -> dict[str, float] | None:
    """在任意嵌套 JSON 中深度搜索可解析为「品种 -> 盈亏」的结构。"""
    if _depth > max_depth:
        return None
    got = _try_extract_from_json(obj)
    if got and len(got) >= 2:
        return got
    if isinstance(obj, dict):
        for v in obj.values():
            r = _deep_search_variety_map(v, max_depth, _depth + 1)
            if r:
                return r
    elif isinstance(obj, list):
        for item in obj:
            r = _deep_search_variety_map(item, max_depth, _depth + 1)
            if r:
                return r
    return None


def _should_skip_json_response_url(url: str) -> bool:
    """跳过明显非业务 JSON 的静态资源，避免吞掉内存；保留本站与业务 API。"""
    u = url.lower()
    if any(u.endswith(ext) for ext in (".woff", ".woff2", ".ttf", ".map")):
        return True
    if "/_nuxt/" in u and any(u.endswith(ext) for ext in (".js", ".css", ".json")):
        # Nuxt 构建产物里部分 chunk 名为 *.json 实为 manifest，可跳过
        if "manifest" in u or "builds" in u or "routes" in u:
            return True
    return False


def _extract_from_echarts_js() -> str:
    """在页面内执行的函数字符串，返回 JSON：{ title, varieties } 或 null。
    兼容多种 ECharts 挂载方式。"""
    return r"""() => {
      try {
        const results = [];

        // ======= 寻找 echarts 引用 =======
        let ec = null;
        if (typeof echarts !== 'undefined') ec = echarts;
        else if (typeof window.echarts !== 'undefined') ec = window.echarts;

        // 尝试从 webpack/vite 模块缓存中找 echarts
        if (!ec) {
          try {
            const mods = window.__LOADED_MODULES__ || window.webpackChunk || [];
            for (const m of Object.values(window)) {
              if (m && typeof m === 'object' && m.getInstanceByDom) { ec = m; break; }
            }
          } catch(e) {}
        }

        // 尝试从带 _echarts_instance_ 属性的 div 反查
        if (!ec) {
          const divs = document.querySelectorAll('[_echarts_instance_]');
          for (const div of divs) {
            // 有些版本在 div 上挂了 __bindChart 或类似属性
            for (const k of Object.keys(div)) {
              if (k.startsWith('__bindChart') || k.startsWith('bindChart')) {
                const chart = div[k];
                if (chart && chart.getOption) {
                  ec = { getInstanceByDom: () => chart };
                  break;
                }
              }
            }
            if (ec) break;
            // 检查 Vue 挂载: __vue__ 或 __vueParentComponent
            const vueComp = div.__vue__ || div.__vueParentComponent;
            if (vueComp) {
              const proxy = vueComp.proxy || vueComp;
              // vue-echarts 组件上通常有 chart 属性
              const chart = proxy.chart || proxy.$refs?.chart;
              if (chart && chart.getOption) {
                ec = { getInstanceByDom: () => chart };
                break;
              }
            }
          }
        }

        // ======= 提取数据 =======
        const xDataOf = (opt) => {
          const xa = opt.xAxis;
          if (!xa) return [];
          const first = Array.isArray(xa) ? xa[0] : xa;
          if (!first || !Array.isArray(first.data)) return [];
          return first.data.map((x) => String(x));
        };

        const extractFromChart = (chart) => {
          let opt;
          try { opt = chart.getOption(); } catch (e) { return null; }
          if (!opt) return null;
          const titleText = (opt.title && opt.title[0] && opt.title[0].text)
            ? String(opt.title[0].text) : '';
          const serList = opt.series;
          if (!Array.isArray(serList) || !serList.length) return null;
          const cats = xDataOf(opt);
          const merged = {};
          for (const ser of serList) {
            const data = ser && ser.data;
            if (!Array.isArray(data) || data.length === 0) continue;
            for (let i = 0; i < data.length; i++) {
              let v = data[i];
              let nm = '';
              if (v && typeof v === 'object' && !Array.isArray(v)) {
                nm = v.name != null ? String(v.name) : '';
                v = v.value;
              }
              if (Array.isArray(v)) v = v[0];
              if (!nm && cats[i]) nm = String(cats[i]);
              if (!nm) continue;
              const num = typeof v === 'number' ? v : parseFloat(String(v).replace(/,/g, ''));
              if (!isNaN(num) && num !== 0) merged[nm] = (merged[nm] || 0) + num;
              else if (!isNaN(num) && !(nm in merged)) merged[nm] = 0;
            }
          }
          if (Object.keys(merged).length < 2) return null;
          return { title: titleText, pairs: merged };
        };

        // 方式 1：通过 _echarts_instance_ 属性找到所有图表容器
        const containers = document.querySelectorAll('[_echarts_instance_]');
        if (ec && containers.length) {
          for (const el of containers) {
            let chart = null;
            try { chart = ec.getInstanceByDom(el); } catch (e) {}
            if (!chart) continue;
            const r = extractFromChart(chart);
            if (r) results.push(r);
          }
        }

        // 方式 2：如果没有 _echarts_instance_，扫描所有 div
        if (results.length === 0 && ec) {
          for (const el of document.querySelectorAll('div')) {
            let chart = null;
            try { chart = ec.getInstanceByDom(el); } catch (e) { continue; }
            if (!chart) continue;
            const r = extractFromChart(chart);
            if (r) results.push(r);
          }
        }

        if (results.length === 0) return null;

        const onlyPnL = results.filter((r) => r.title && r.title.includes('品种盈亏'));
        const pool = onlyPnL.length ? onlyPnL : results;
        const score = (r) => {
          let s = Object.keys(r.pairs).length;
          const t = r.title || '';
          if (t.includes('品种盈亏')) s += 100000;
          else if (t.includes('品种')) s += 1000;
          return s;
        };
        pool.sort((a, b) => score(b) - score(a));
        const pick = pool[0];
        if (!pick) return null;
        return JSON.stringify({ title: pick.title || '', varieties: pick.pairs });
      } catch (e) {
        return null;
      }
    }"""


def _decode_echarts_json(raw: str) -> tuple[dict[str, float] | None, str]:
    """解析 ECharts 脚本返回值：{ title, varieties } 或旧版扁平 { 名: 数 }。"""
    try:
        m = json.loads(raw)
    except Exception:
        return None, ""
    if not isinstance(m, dict):
        return None, ""
    title = ""
    if "varieties" in m and isinstance(m["varieties"], dict):
        title = str(m.get("title") or "")
        out: dict[str, float] = {}
        for k, v in m["varieties"].items():
            n = _parse_number(v)
            if n is not None:
                out[str(k)] = float(n)
        return (out if len(out) >= 2 else None), title
    out2: dict[str, float] = {}
    for k, v in m.items():
        if k in ("title", "varieties"):
            continue
        n = _parse_number(v)
        if n is not None:
            out2[str(k)] = float(n)
    return (out2 if len(out2) >= 2 else None), ""


def _extract_echarts_from_all_frames(page, echarts_js: str) -> tuple[dict[str, float] | None, str]:
    """在所有 frame 中读取 ECharts；返回 (品种映射, 图表标题)。"""
    best: dict[str, float] | None = None
    best_title = ""
    best_url = ""
    best_score = -1.0

    def frame_score(u: str, title: str, data: dict[str, float]) -> float:
        sc = _variety_map_quality_score(data)
        if "品种盈亏" in (title or ""):
            sc += 8000.0
        ul = u.lower()
        if "dpswang.com" in ul:
            sc += 400.0
        if "qiweihu" in ul:
            sc -= 200.0
        return sc

    for fr in page.frames:
        try:
            u = fr.url or ""
        except Exception:
            u = ""
        if not u or u.startswith("about:"):
            continue
        try:
            raw = fr.evaluate(echarts_js)
        except Exception:
            continue
        if not isinstance(raw, str) or not raw.strip():
            continue
        m, title = _decode_echarts_json(raw)
        if not m:
            continue
        sc = frame_score(u, title, m)
        if sc > best_score:
            best_score = sc
            best, best_title, best_url = m, title, u
        elif best is not None and abs(sc - best_score) < 1e-6 and len(m) > len(best):
            best, best_title, best_url = m, title, u

    return (best if best and len(best) >= 2 else None), (best_title or "")


def _extract_from_iframe_content(page) -> dict[str, float] | None:
    """从 iframe 中尝试多种方式提取品种盈亏数据（包括七尾狐 iframe 内的 ECharts）。"""
    best: dict[str, float] | None = None
    # 在每个 frame 中尝试更激进的提取方式
    extract_js = r"""() => {
      try {
        // 尝试从 Vue 组件的 data 中提取（七尾狐可能用 Vue）
        const findInVue = () => {
          const app = document.querySelector('#app') || document.querySelector('[data-v-app]');
          if (!app || !app.__vue_app__) return null;
          const inst = app.__vue_app__._instance;
          if (!inst) return null;
          const findData = (obj, depth) => {
            if (depth > 8 || !obj) return null;
            if (obj.setupState || obj.data) {
              const state = obj.setupState || obj.data;
              const keys = Object.keys(state);
              for (const k of keys) {
                const val = state[k];
                if (Array.isArray(val) && val.length > 5) {
                  const first = val[0];
                  if (first && typeof first === 'object' && (first.name || first.varietyName)) {
                    return val;
                  }
                }
              }
            }
            if (obj.subTree && obj.subTree.component) {
              return findData(obj.subTree.component, depth + 1);
            }
            if (obj.children) {
              for (const ch of (Array.isArray(obj.children) ? obj.children : [])) {
                if (ch && ch.component) {
                  const r = findData(ch.component, depth + 1);
                  if (r) return r;
                }
              }
            }
            return null;
          };
          return findData(inst, 0);
        };

        // 尝试从全局变量提取
        const findInGlobal = () => {
          for (const k of Object.keys(window)) {
            if (k.startsWith('_') || k === 'location') continue;
            try {
              const v = window[k];
              if (Array.isArray(v) && v.length > 5 && v[0] && typeof v[0] === 'object') {
                if (v[0].name || v[0].varietyName || v[0].productName) return v;
              }
            } catch (e) { continue; }
          }
          return null;
        };

        const arr = findInVue() || findInGlobal();
        if (arr && Array.isArray(arr)) {
          const out = {};
          for (const item of arr) {
            const nm = item.name || item.varietyName || item.productName || '';
            const val = item.profit || item.pnl || item.value || item.profitLoss || 0;
            if (nm) out[String(nm)] = typeof val === 'number' ? val : parseFloat(String(val)) || 0;
          }
          if (Object.keys(out).length >= 2) return JSON.stringify(out);
        }
        return null;
      } catch (e) { return null; }
    }"""
    for fr in page.frames:
        try:
            u = fr.url or ""
        except Exception:
            continue
        if not u or u.startswith("about:") or "login" in u.lower():
            continue
        try:
            raw = fr.evaluate(extract_js)
        except Exception:
            continue
        if not isinstance(raw, str) or not raw.strip():
            continue
        try:
            m = json.loads(raw)
        except Exception:
            continue
        if isinstance(m, dict) and len(m) >= 2:
            if best is None or len(m) > len(best):
                best = m
    return best


def _extract_from_nuxt_and_inline_scripts(page) -> dict[str, float] | None:
    """从主文档 window.__NUXT__ / __NUXT_DATA__ / application/json 脚本中深度搜索品种盈亏。"""
    best: dict[str, float] | None = None

    def _is_valid_variety_map(d: dict[str, float] | None) -> bool:
        if not d or len(d) < 2:
            return False
        # 必须至少有一个中文键（品种名一定是中文）
        return any(re.search(r"[\u4e00-\u9fff]", k) for k in d)

    for expr in ("() => window.__NUXT__ ?? null", "() => window.__NUXT_DATA__ ?? null"):
        try:
            data = page.evaluate(expr)
        except Exception:
            continue
        if data is None:
            continue
        got = _deep_search_variety_map(data)
        if _is_valid_variety_map(got) and (best is None or len(got) > len(best)):
            best = got
    try:
        texts = page.evaluate(
            """() => {
              const out = [];
              for (const s of document.querySelectorAll(
                'script[type="application/json"], script#__NUXT_DATA__'
              )) {
                const t = s.textContent;
                if (t && t.length > 30 && t.length < 6000000) out.push(t);
              }
              return out;
            }"""
        )
        if isinstance(texts, list):
            for raw in texts[:40]:
                if not isinstance(raw, str):
                    continue
                try:
                    obj = json.loads(raw)
                except Exception:
                    continue
                got = _deep_search_variety_map(obj)
                if _is_valid_variety_map(got) and (best is None or len(got) > len(best)):
                    best = got
    except Exception:
        pass
    return best


def _click_yingkui_fenxi_tab(page) -> None:
    """点击主内容区「盈亏分析」Tab（截图中与累计净值、历史权益同一排，第3个标签）。"""

    # 辅助：打印元素详细信息
    def _print_element_info(loc, label: str) -> None:
        try:
            tag = loc.evaluate("el => el.tagName")
            text = loc.evaluate("el => el.textContent.trim().slice(0, 40)")
            parent_text = loc.evaluate(
                "el => el.parentElement ? el.parentElement.textContent.trim().slice(0, 80) : ''"
            )
            print(f"  [调试] {label}: <{tag}> text=\"{text}\" parent_text=\"{parent_text}\"")
        except Exception:
            print(f"  [调试] {label}: 无法获取元素信息")

    # 策略1: 精确文本匹配，选最小的（叶子）元素
    # 使用 :text-is() 只匹配 own text content 完全等于目标的元素（不包含子元素文字拼合）
    leaf_selectors = [
        ':text-is("盈亏分析")',          # 精确文本的叶子节点
        'a:text-is("盈亏分析")',
        'span:text-is("盈亏分析")',
        'div:text-is("盈亏分析")',
    ]
    for sel in leaf_selectors:
        try:
            loc = page.locator(sel)
            cnt = loc.count()
            if cnt == 0:
                continue
            # 如果有多个匹配，打印所有并选第一个
            print(f"  [调试] selector \"{sel}\" 匹配到 {cnt} 个元素")
            for i in range(min(cnt, 5)):
                _print_element_info(loc.nth(i), f"  匹配[{i}]")
            # 点第一个
            loc.first.scroll_into_view_if_needed(timeout=5000)
            loc.first.click(timeout=8000)
            page.wait_for_timeout(3000)
            print(f"  [调试] 已点击 selector \"{sel}\" 的第一个元素")
            return
        except Exception as e:
            print(f"  [调试] selector \"{sel}\" 失败: {e}")
            continue

    # 策略2: 查找页面中所有 Tab 样式元素，按照截图顺序找到第3个
    # 截图中 Tab 顺序：累计净值、历史权益、盈亏分析(第3个)、仓位分析、交易周期...
    print("  [调试] 叶子选择器均未命中，尝试按 Tab 顺序定位...")
    tab_bar_selectors = [
        '[class*="tab"] a, [class*="tab"] span, [class*="tab"] div',
        '[class*="nav"] a, [class*="nav"] span',
        '[role="tablist"] [role="tab"]',
    ]
    for bar_sel in tab_bar_selectors:
        try:
            tabs = page.locator(bar_sel)
            cnt = tabs.count()
            if cnt < 3:
                continue
            print(f"  [调试] Tab 栏 selector \"{bar_sel}\" 找到 {cnt} 个元素:")
            for i in range(min(cnt, 12)):
                try:
                    txt = tabs.nth(i).evaluate("el => el.textContent.trim().slice(0, 20)")
                    print(f"    [{i}] \"{txt}\"")
                    if "盈亏分析" in txt and len(txt) <= 6:
                        tabs.nth(i).scroll_into_view_if_needed(timeout=5000)
                        tabs.nth(i).click(timeout=8000)
                        page.wait_for_timeout(3000)
                        print(f"  [调试] 已点击 Tab[{i}] \"{txt}\"")
                        return
                except Exception:
                    continue
        except Exception:
            continue

    # 策略3: get_by_text（最后手段）
    try:
        loc = page.get_by_text("盈亏分析", exact=True)
        cnt = loc.count()
        print(f"  [调试] get_by_text(exact=True) 匹配到 {cnt} 个")
        if cnt > 0:
            for i in range(min(cnt, 5)):
                _print_element_info(loc.nth(i), f"  get_by_text[{i}]")
            loc.first.scroll_into_view_if_needed(timeout=5000)
            loc.first.click(timeout=8000)
            page.wait_for_timeout(3000)
            print("  [调试] 已点击 get_by_text 第一个元素")
            return
    except Exception:
        pass

    print("  [警告] 未能找到并点击「盈亏分析」Tab")


def _login_dpswang(
    page,
    username: str,
    password: str,
    login_url: str,
    timeout_ms: int,
) -> None:
    """在大盘手网登录页填写账号密码并提交（遇验证码/短信验证需改用 storage_state 手动登录）。"""
    page.goto(login_url, wait_until="domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(800)

    try:
        tab = page.get_by_text("账号密码登录", exact=False)
        if tab.count() > 0:
            tab.first.click(timeout=5000)
            page.wait_for_timeout(500)
    except Exception:
        pass

    user_filled = False
    for hint in ("手机号", "手机", "账号", "用户名", "邮箱"):
        try:
            loc = page.get_by_placeholder(hint, exact=False)
            if loc.count() > 0:
                loc.first.fill(username, timeout=8000)
                user_filled = True
                break
        except Exception:
            continue

    if not user_filled:
        try:
            loc = page.locator('input[type="tel"], input[type="text"]').first
            loc.wait_for(state="visible", timeout=8000)
            loc.fill(username, timeout=8000)
            user_filled = True
        except Exception as exc:
            raise RuntimeError(
                "无法定位账号输入框；网站结构可能已变更，或当前为验证码登录页。"
                "请使用 --headed 手动登录并用 --save-storage-state 保存会话。"
            ) from exc

    try:
        page.locator('input[type="password"]').first.fill(password, timeout=8000)
    except Exception as exc:
        raise RuntimeError("无法定位密码输入框。") from exc

    try:
        page.get_by_role("button", name=re.compile("登\\s*录")).first.click(timeout=8000)
    except Exception:
        try:
            page.get_by_text("登录", exact=False).first.click(timeout=8000)
        except Exception as exc:
            raise RuntimeError("无法点击登录按钮。") from exc

    try:
        page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 60_000))
    except Exception:
        pass
    page.wait_for_timeout(1500)

    cur = page.url or ""
    if "login" in cur.lower():
        raise RuntimeError(
            "登录后仍停留在登录页：可能密码错误、需要图形验证码/短信验证，或站点风控。"
            "请改用 --headed 手动完成验证后使用 --save-storage-state 保存 cookies。"
        )


def _dump_api_data(body: Any, max_depth: int = 3, prefix: str = "    ") -> None:
    """打印 API 响应体的结构摘要，便于调试数据格式。"""
    def _dump(obj, depth, pfx):
        if depth > max_depth:
            print(f"{pfx}...")
            return
        if isinstance(obj, dict):
            print(f"{pfx}dict keys={list(obj.keys())[:12]}")
            for k, v in list(obj.items())[:5]:
                if isinstance(v, (dict, list)):
                    print(f"{pfx}  [{k}]:")
                    _dump(v, depth + 1, pfx + "    ")
                else:
                    print(f"{pfx}  [{k}]: {repr(v)[:80]}")
        elif isinstance(obj, list):
            print(f"{pfx}list[{len(obj)}]")
            if obj:
                print(f"{pfx}  [0]:")
                _dump(obj[0], depth + 1, pfx + "    ")
                if len(obj) > 1:
                    print(f"{pfx}  [1]:")
                    _dump(obj[1], depth + 1, pfx + "    ")
        else:
            print(f"{pfx}{repr(obj)[:100]}")
    _dump(body, 0, prefix)


def fetch_variety_pnl_map(
    account_url: str,
    headed: bool = False,
    timeout_ms: int = 90_000,
    *,
    login_url: str = DEFAULT_LOGIN_URL,
    username: str | None = None,
    password: str | None = None,
    storage_state_path: Path | None = None,
    save_storage_state_path: Path | None = None,
) -> dict[str, float]:
    from playwright.sync_api import sync_playwright

    json_snapshots: list[tuple[str, Any]] = []
    all_response_urls: list[str] = []
    seen_urls: set[str] = set()

    def on_response(resp) -> None:
        try:
            u = resp.url
            all_response_urls.append(u)
            if u in seen_urls:
                return
            if resp.status >= 400:
                return
            if _should_skip_json_response_url(u):
                return
            ct = (resp.headers or {}).get("content-type", "") or ""
            is_json_ct = "json" in ct.lower()
            is_json_url = u.lower().split("?", 1)[0].endswith(".json")
            # 对 qiweihu.cn 域名的请求也尝试解析（可能不声明 json content-type）
            is_qiweihu = "qiweihu" in u.lower()
            is_dpswapi = "dpswapi" in u.lower()
            if not (is_json_ct or is_json_url or is_qiweihu or is_dpswapi):
                return
            body = resp.json()
            seen_urls.add(u)
            json_snapshots.append((u, body))
        except Exception:
            return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headed)
        ctx_kw: dict[str, Any] = {
            "locale": "zh-CN",
            "user_agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        }
        if storage_state_path is not None and storage_state_path.is_file():
            ctx_kw["storage_state"] = str(storage_state_path)

        context = browser.new_context(**ctx_kw)
        page = context.new_page()
        page.on("response", on_response)

        use_password_login = bool(
            username and password is not None and str(password).strip() != ""
        )
        if use_password_login and "storage_state" not in ctx_kw:
            _login_dpswang(page, username.strip(), password, login_url, timeout_ms)

        page.goto(account_url, wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(3000)

        # 早期检测：如果被重定向到登录页
        current_url = page.url or ""
        if "/login" in current_url.lower():
            if headed:
                # 有头模式：等待用户手动登录
                print("  [登录] Cookie 已过期，请在浏览器中手动登录...")
                print("  [登录] 登录成功后页面会自动跳转，脚本将继续执行。")
                try:
                    page.wait_for_url(
                        lambda u: "/login" not in u.lower(),
                        timeout=300_000,  # 最长等 5 分钟
                    )
                    print("  [登录] ✓ 登录成功，继续执行...")
                    page.wait_for_timeout(2000)
                    # 登录后重新访问目标页面
                    page.goto(account_url, wait_until="domcontentloaded", timeout=timeout_ms)
                    page.wait_for_timeout(3000)
                except Exception:
                    if save_storage_state_path is not None:
                        save_storage_state_path.parent.mkdir(parents=True, exist_ok=True)
                        context.storage_state(path=str(save_storage_state_path))
                    browser.close()
                    raise RuntimeError("等待登录超时（5分钟），请重试。")
            else:
                # 无头模式：直接报错
                browser.close()
                raise RuntimeError(
                    "Cookie 已过期，页面被重定向到登录页。"
                    "请用 --head --save-storage-state dpswang_auth.json 重新手动登录。"
                )

        print("  [步骤1] 页面已加载，准备点击「盈亏分析」Tab...")

        # 点击「盈亏分析」
        _click_yingkui_fenxi_tab(page)

        # 滚动以触发懒加载图表
        print("  [步骤2] 滚动页面触发懒加载...")
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(2000)
            page.evaluate("window.scrollTo(0, 0)")
            page.wait_for_timeout(1000)
        except Exception:
            pass

        # 等待「品种盈亏」文案出现
        print("  [步骤3] 等待「品种盈亏」图表加载...")
        pnl_visible = False
        try:
            page.get_by_text("品种盈亏", exact=False).first.wait_for(
                state="visible", timeout=20_000
            )
            pnl_visible = True
            print("  [步骤3] ✓ 已看到「品种盈亏」文字")
        except Exception:
            print("  [步骤3] ✗ 未找到「品种盈亏」文字，可能 Tab 未切换或数据未加载")

        # 给图表渲染充分时间
        page.wait_for_timeout(5000)

        # ============ 提取数据 ============
        print("  [步骤4] 开始从页面提取数据...")

        # 方法 A: ECharts 实例
        echarts_js = _extract_from_echarts_js()
        echarts_map, echarts_title = _extract_echarts_from_all_frames(page, echarts_js)
        if echarts_map:
            print(f"  [步骤4a] ECharts 提取成功: {len(echarts_map)} 个品种, 标题=\"{echarts_title}\"")
        else:
            print("  [步骤4a] ECharts 提取未返回数据")

        # 方法 B: Nuxt / 内联 JSON
        nuxt_map = _extract_from_nuxt_and_inline_scripts(page)
        if nuxt_map:
            print(f"  [步骤4b] Nuxt/内联脚本提取成功: {len(nuxt_map)} 个品种")
        else:
            print("  [步骤4b] Nuxt/内联脚本未返回数据")

        # 方法 C: 网络 JSON 响应 — 优先 varietyProfit API
        parsed_from_api: dict[str, float] | None = None
        best_api_url = ""
        variety_profit_data: dict[str, float] | None = None

        for api_url, snap in json_snapshots:
            # 专门处理品种盈亏 API（最高优先级）
            if "varietyProfit" in api_url or "variety_profit" in api_url or "variety-profit" in api_url:
                print(f"  [步骤4c] 发现品种盈亏 API: {api_url}")
                got = _try_extract_from_json(snap) or _deep_search_variety_map(snap)
                if got:
                    variety_profit_data = got
                    print(f"  [步骤4c] varietyProfit 解析成功: {len(got)} 个品种")
                    preview = list(got.items())[:5]
                    for k, v in preview:
                        print(f"    {k}: {v}")
                else:
                    print(f"  [步骤4c] varietyProfit 标准解析失败，尝试 dump...")
                    _dump_api_data(snap)
                continue

            got = _try_extract_from_json(snap) or _deep_search_variety_map(snap)
            if not got:
                continue
            if parsed_from_api is None or len(got) > len(parsed_from_api):
                parsed_from_api, best_api_url = got, api_url

        # varietyProfit 优先于其他任何 API
        if variety_profit_data:
            parsed_from_api = variety_profit_data
            best_api_url = "varietyProfit"

        if parsed_from_api:
            print(f"  [步骤4c] 最终选择: {len(parsed_from_api)} 个品种 (来源: {best_api_url[:80]})")
        else:
            print(f"  [步骤4c] 网络 API 未返回品种数据 (共拦截 {len(json_snapshots)} 条 JSON 响应)")

        # 方法 D: 从 iframe 的 contentWindow 直接读取（七尾狐图表）
        iframe_map = _extract_from_iframe_content(page)
        if iframe_map:
            print(f"  [步骤4d] iframe 内容提取成功: {len(iframe_map)} 个品种")
        else:
            print("  [步骤4d] iframe 内容未返回数据")

        # 诊断：始终打印关键信息
        print(f"  [诊断] 页面 frame 数: {len(page.frames)}")
        for fr in page.frames:
            u = fr.url or ""
            print(f"    frame: {u[:120]}")
        print(f"  [诊断] 拦截到 {len(json_snapshots)} 条 JSON 响应:")
        for url, body in json_snapshots:
            body_type = type(body).__name__
            body_preview = ""
            if isinstance(body, dict):
                keys = list(body.keys())[:8]
                body_preview = f"keys={keys}"
            elif isinstance(body, list):
                body_preview = f"list[{len(body)}]"
                if body and isinstance(body[0], dict):
                    body_preview += f" first_keys={list(body[0].keys())[:6]}"
            print(f"    {url[:100]} => {body_type}: {body_preview}")
        qwh_urls = [u for u in all_response_urls if "qiweihu" in u.lower()]
        if qwh_urls:
            print(f"  [诊断] 七尾狐相关 URL ({len(qwh_urls)} 条):")
            for u in qwh_urls[:15]:
                print(f"    {u[:150]}")

        if save_storage_state_path is not None:
            save_storage_state_path.parent.mkdir(parents=True, exist_ok=True)
            context.storage_state(path=str(save_storage_state_path))

        browser.close()

    result = _pick_best_variety_maps(
        [
            ("echarts", echarts_map, echarts_title),
            ("iframe", iframe_map, "品种盈亏"),
            ("api", parsed_from_api, ""),
            ("nuxt", nuxt_map, ""),
        ]
    )
    if not result:
        raise RuntimeError(
            "未能解析到「品种盈亏」数据。请确认：\n"
            "  1. 用 --headed 运行后能看到「盈亏分析」Tab 已高亮且下方有「品种盈亏」柱状图\n"
            "  2. 若图表区域显示的是七尾狐登录页，说明该数据需要七尾狐授权\n"
            "  3. 请把上面的 [诊断] 信息贴出来以便进一步排查"
        )
    return result


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _read_header_varieties(ws) -> list[str]:
    if ws.max_row < 1:
        return []
    row1 = [ws.cell(row=1, column=c).value for c in range(2, ws.max_column + 1)]
    out = []
    for v in row1:
        if v is None or str(v).strip() == "":
            continue
        out.append(str(v).strip())
    return out


def _read_data_row_as_map(ws, data_row: int) -> dict[str, float]:
    """按第 1 行表头读取指定数据行，避免列顺序变化导致错位。"""
    out: dict[str, float] = {}
    for c in range(2, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None or str(h).strip() == "":
            continue
        name = str(h).strip()
        cell = ws.cell(row=data_row, column=c).value
        if cell is None or str(cell).strip() == "":
            continue
        try:
            out[name] = float(cell)
        except (TypeError, ValueError):
            continue
    return out


def _write_daily_and_diff(
    excel_path: Path,
    today: date,
    variety_to_value: dict[str, float],
    user_name: str = "",
) -> None:
    """将品种盈亏数据写入 Excel。每个用户有两个 Sheet：'用户名_每日数据' 和 '用户名_每日变化'。"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    sheet_daily = f"{user_name}_{SHEET_DAILY_SUFFIX}" if user_name else SHEET_DAILY_SUFFIX
    sheet_diff = f"{user_name}_{SHEET_DIFF_SUFFIX}" if user_name else SHEET_DIFF_SUFFIX

    ws_d = _ensure_sheet(wb, sheet_daily)
    ws_f = _ensure_sheet(wb, sheet_diff)

    # 合并表头：已有品种 ∪ 今日新品种
    existing = _read_header_varieties(ws_d)
    all_varieties = sorted(set(existing) | set(variety_to_value.keys()), key=lambda x: x)

    # 写表头
    ws_d.cell(row=1, column=1, value="日期")
    ws_f.cell(row=1, column=1, value="日期")
    for i, name in enumerate(all_varieties, start=2):
        ws_d.cell(row=1, column=i, value=name)
        ws_f.cell(row=1, column=i, value=name)

    # 找是否已有今日行（允许同一天重跑覆盖）
    today_str = today.isoformat()
    data_row_idx = None
    for r in range(2, ws_d.max_row + 1):
        v = ws_d.cell(row=r, column=1).value
        if v is not None and str(v).strip() == today_str:
            data_row_idx = r
            break

    if data_row_idx is None:
        data_row_idx = max(2, ws_d.max_row + 1)

    ws_d.cell(row=data_row_idx, column=1, value=today_str)
    for col, name in enumerate(all_varieties, start=2):
        val = variety_to_value.get(name)
        ws_d.cell(row=data_row_idx, column=col, value=val)

    # 前一有效数据行（按日期排序找前一日）
    prev_row_map: dict[str, float] = {}
    prev_date: str | None = None
    rows_meta: list[tuple[str, int]] = []
    for r in range(2, ws_d.max_row + 1):
        ds = ws_d.cell(row=r, column=1).value
        if ds is None:
            continue
        rows_meta.append((str(ds).strip(), r))
    rows_meta.sort(key=lambda x: x[0])
    for i, (ds, r) in enumerate(rows_meta):
        if ds == today_str and i > 0:
            _pdate, prev_r = rows_meta[i - 1]
            prev_date = _pdate
            prev_row_map = _read_data_row_as_map(ws_d, prev_r)
            break

    # 写变化表：仅当有前一日数据时写入一行
    if prev_date is not None and prev_row_map:
        diff_row_idx = None
        for r in range(2, ws_f.max_row + 1):
            v = ws_f.cell(row=r, column=1).value
            if v is not None and str(v).strip() == today_str:
                diff_row_idx = r
                break
        if diff_row_idx is None:
            diff_row_idx = max(2, ws_f.max_row + 1)

        ws_f.cell(row=diff_row_idx, column=1, value=today_str)
        for col, name in enumerate(all_varieties, start=2):
            cur = variety_to_value.get(name)
            pr = prev_row_map.get(name)
            if cur is None and pr is None:
                ws_f.cell(row=diff_row_idx, column=col, value=None)
            elif cur is None:
                ws_f.cell(row=diff_row_idx, column=col, value=-float(pr))
            elif pr is None:
                ws_f.cell(row=diff_row_idx, column=col, value=float(cur))
            else:
                ws_f.cell(row=diff_row_idx, column=col, value=float(cur) - float(pr))

    # 冻结首行
    ws_d.freeze_panes = "A2"
    ws_f.freeze_panes = "A2"

    wb.save(excel_path)


def _do_login_only(save_path: Path, login_url: str, timeout_ms: int) -> int:
    """纯登录模式：打开登录页，等待手动登录，保存 Cookie 后退出。"""
    from playwright.sync_api import sync_playwright

    print("[登录模式] 打开浏览器，请在页面中完成登录...")
    print(f"[登录模式] 登录成功后 Cookie 将保存到: {save_path}")
    print(f"[登录模式] 最长等待 5 分钟，登录完成后自动保存并关闭。\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            locale="zh-CN",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.goto(login_url, wait_until="domcontentloaded", timeout=timeout_ms)

        # 等待用户登录完成（URL 离开 /login）
        try:
            page.wait_for_url(
                lambda u: "/login" not in u.lower(),
                timeout=300_000,
            )
        except Exception:
            print("[登录模式] ✗ 等待超时（5分钟），未检测到登录成功。", file=sys.stderr)
            browser.close()
            return 1

        page.wait_for_timeout(2000)
        save_path.parent.mkdir(parents=True, exist_ok=True)
        context.storage_state(path=str(save_path))
        browser.close()

    print(f"\n[登录模式] ✓ 登录成功，Cookie 已保存到: {save_path}")
    print("[登录模式] 之后运行数据采集只需：")
    print(f"  python fetch_variety_pnl.py --storage-state {save_path.name}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="采集大盘手网品种盈亏并写入 Excel（自动遍历所有配置用户）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "两种用法：\n"
            "  1. 登录保存 Cookie（仅登录，不抓数据）：\n"
            "     python fetch_variety_pnl.py --save-storage-state dpswang_auth.json\n\n"
            "  2. 抓取数据（使用已保存的 Cookie）：\n"
            "     python fetch_variety_pnl.py --storage-state dpswang_auth.json\n"
        ),
    )
    parser.add_argument(
        "--login-url",
        default=DEFAULT_LOGIN_URL,
        help="登录页 URL",
    )
    parser.add_argument(
        "--username",
        default=None,
        help=f"登录账号（默认读环境变量 {ENV_USERNAME}）",
    )
    parser.add_argument(
        "--password",
        default=None,
        help=f"登录密码（默认读环境变量 {ENV_PASSWORD}）",
    )
    parser.add_argument(
        "--storage-state",
        default=None,
        help="加载已保存的 Cookie 文件（用于数据采集）",
    )
    parser.add_argument(
        "--save-storage-state",
        default=None,
        help="保存登录 Cookie 到该路径（不带 --storage-state 时为纯登录模式）",
    )
    parser.add_argument(
        "--excel",
        default=str(_script_dir() / DEFAULT_EXCEL_NAME),
        help="输出 Excel 路径",
    )
    parser.add_argument("--headed", "--head", action="store_true", help="有头模式，便于调试")
    parser.add_argument("--timeout", type=int, default=90_000, help="页面超时（毫秒）")
    args = parser.parse_args()

    save_storage_path = Path(args.save_storage_state).resolve() if args.save_storage_state else None

    # ===== 纯登录模式：只有 --save-storage-state，没有 --storage-state =====
    if save_storage_path and not args.storage_state:
        return _do_login_only(save_storage_path, args.login_url, args.timeout)

    # ===== 数据采集模式 =====
    login_user = (args.username or os.environ.get(ENV_USERNAME) or "").strip()
    pwd = args.password or os.environ.get(ENV_PASSWORD) or ""
    storage_path = Path(args.storage_state).resolve() if args.storage_state else None

    today = date.today()
    out_path = Path(args.excel)
    success_count = 0
    fail_count = 0

    for user_name, account_id in ACCOUNTS.items():
        account_url = DEFAULT_ACCOUNT_URL.format(account_id=account_id.strip())
        print(f"\n{'='*60}")
        print(f"[用户] {user_name} (账户ID: {account_id.strip()}, URL: {account_url})")
        print(f"{'='*60}")

        try:
            data = fetch_variety_pnl_map(
                account_url,
                headed=args.headed,
                timeout_ms=args.timeout,
                login_url=args.login_url,
                username=login_user if login_user else None,
                password=pwd if pwd else None,
                storage_state_path=storage_path,
                save_storage_state_path=save_storage_path,
            )
        except Exception as e:
            print(f"  抓取失败: {e}", file=sys.stderr)
            fail_count += 1
            continue

        try:
            _write_daily_and_diff(out_path, today, data, user_name=user_name)
        except Exception as e:
            print(f"  写入 Excel 失败: {e}", file=sys.stderr)
            fail_count += 1
            continue

        success_count += 1
        print(f"  ✓ 已写入，日期 {today.isoformat()} ，品种数 {len(data)}")
        preview = list(data.items())[:3]
        for k, v in preview:
            print(f"    {k}: {v}")
        if len(data) > 3:
            print("    ...")

    print(f"\n{'='*60}")
    print(f"[汇总] 成功 {success_count} / 失败 {fail_count} / 总计 {len(ACCOUNTS)}")
    print(f"[输出] {out_path}")
    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
